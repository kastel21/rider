"""CSV / Excel export helpers for M&E matrix-style table dicts."""

from __future__ import annotations

import csv
import zipfile
from io import BytesIO, StringIO
from typing import Any

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ..selectors import sunday_of_week
from .me_accidents_data import load_me_accidents_incomplete_data


def matrix_table_to_rows(table: dict[str, Any]) -> tuple[list[str], list[list[str]]]:
    headers = [str(c.get("label") or "").replace("\r\n", " ").replace("\n", " ") for c in table["columns"]]
    body: list[list[str]] = []
    for row in table["rows"]:
        body.append([str(cell.get("text") or "") for cell in row])
    return headers, body


def _csv_bytes(headers: list[str], rows: list[list[str]]) -> bytes:
    buf = StringIO()
    w = csv.writer(buf, lineterminator="\r\n")
    w.writerow(headers)
    w.writerows(rows)
    return buf.getvalue().encode("utf-8-sig")


def _xlsx_bytes(headers: list[str], rows: list[list[str]], *, sheet_title: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] if sheet_title else "Data"
    ws.append(headers)
    for r in rows:
        ws.append(r)
    for i, _ in enumerate(headers, start=1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = min(48, max(10, len(headers[i - 1]) + 2))
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def filename_safe(s: str) -> str:
    return "".join(c if c.isalnum() or c in "-_" else "_" for c in s)[:180]


def response_matrix_csv(*, table: dict[str, Any], download_stem: str) -> HttpResponse:
    headers, rows = matrix_table_to_rows(table)
    data = _csv_bytes(headers, rows)
    resp = HttpResponse(data, content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="{filename_safe(download_stem)}.csv"'
    return resp


def response_matrix_xlsx(*, table: dict[str, Any], download_stem: str, sheet_title: str) -> HttpResponse:
    headers, rows = matrix_table_to_rows(table)
    data = _xlsx_bytes(headers, rows, sheet_title=sheet_title)
    resp = HttpResponse(
        data,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename_safe(download_stem)}.xlsx"'
    return resp


def _accidents_district_rows(stats: list[dict]) -> tuple[list[str], list[list[str]]]:
    headers = ["Province", "District", "Rider accidents"]
    rows = [[str(r["province_name"]), str(r["district_name"]), str(r["rider_accidents"])] for r in stats]
    return headers, rows


def _accidents_detail_rows(details: list[dict]) -> tuple[list[str], list[list[str]]]:
    headers = ["Rider", "Bike", "Cause of accident", "Status of bike", "Status of rider"]
    rows = [
        [
            str(d["rider_display"]),
            str(d["bike_code"]),
            str(d["accident_cause"] or ""),
            str(d["bike_status_display"]),
            str(d["rider_injury_display"]),
        ]
        for d in details
    ]
    return headers, rows


def _accidents_incomplete_rows(stats: list[dict]) -> tuple[list[str], list[list[str]]]:
    headers = [
        "Province",
        "District",
        "Incomplete bike transport trips",
        "Non-IST total",
        "Ambulance",
        "Alternative IP transport",
        "MoHCC arranged transport",
        "Courier",
        "Other non-IST",
        "Comments",
    ]
    rows = []
    for r in stats:
        rows.append(
            [
                str(r["province_name"]),
                str(r["district_name"]),
                str(r["incomplete_bike_transport_trips"]),
                str(r["specimens_non_ist_total"]),
                str(r["specimens_ambulance"]),
                str(r["specimens_alternative_ip_transport"]),
                str(r["specimens_mohcc_arranged_transport"]),
                str(r["specimens_courier"]),
                str(r["specimens_other_non_ist"]),
                str(r["comments"] or ""),
            ]
        )
    return headers, rows


def response_accidents_zip_csv(*, week_start, bundle: dict[str, list]) -> HttpResponse:
    stats = bundle["transport_stats"]
    details = bundle["accident_details"]
    parts = [
        ("district_accidents.csv", *_accidents_district_rows(stats)),
        ("accident_details.csv", *_accidents_detail_rows(details)),
        ("incomplete_trips.csv", *_accidents_incomplete_rows(stats)),
    ]
    bio = BytesIO()
    with zipfile.ZipFile(bio, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, hdr, body in parts:
            zf.writestr(name, _csv_bytes(hdr, body), compress_type=zipfile.ZIP_DEFLATED)
    stem = f"me-accidents-incomplete-{week_start.isoformat()}"
    resp = HttpResponse(bio.getvalue(), content_type="application/zip")
    resp["Content-Disposition"] = f'attachment; filename="{filename_safe(stem)}.zip"'
    return resp


def response_accidents_xlsx(*, week_start, bundle: dict[str, list]) -> HttpResponse:
    stats = bundle["transport_stats"]
    details = bundle["accident_details"]
    wb = Workbook()
    # Sheet 1
    ws1 = wb.active
    ws1.title = "District accidents"
    h1, r1 = _accidents_district_rows(stats)
    ws1.append(h1)
    for row in r1:
        ws1.append(row)
    ws2 = wb.create_sheet("Accident details")
    h2, r2 = _accidents_detail_rows(details)
    ws2.append(h2)
    for row in r2:
        ws2.append(row)
    ws3 = wb.create_sheet("Incomplete trips")
    h3, r3 = _accidents_incomplete_rows(stats)
    ws3.append(h3)
    for row in r3:
        ws3.append(row)
    bio = BytesIO()
    wb.save(bio)
    stem = f"me-accidents-incomplete-{week_start.isoformat()}"
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename_safe(stem)}.xlsx"'
    return resp


def matrix_download_stem(table: dict[str, Any], *, prefix: str) -> str:
    ws = table["window_start"]
    we = table["window_end"]
    return f"{prefix}-{ws.isoformat()}_to_{sunday_of_week(we).isoformat()}"


def week_download_stem(*, week_start, prefix: str) -> str:
    end = sunday_of_week(week_start)
    return f"{prefix}-{week_start.isoformat()}_to_{end.isoformat()}"


def overview_download_stem(me_metrics: dict[str, Any]) -> str:
    ws = me_metrics["window_start"]
    we = me_metrics["window_end"]
    return f"me-overview-{ws.isoformat()}_to_{sunday_of_week(we).isoformat()}"


def _rows_from_dicts(rows: list[dict], *keys: str) -> list[list[str]]:
    return [[str(r.get(k, "")) for k in keys] for r in rows]


def build_me_overview_export_parts(me_metrics: dict[str, Any]) -> list[tuple[str, list[str], list[list[str]]]]:
    """Named tables on the M&E overview page for zip / multi-sheet export."""
    parts: list[tuple[str, list[str], list[list[str]]]] = []

    parts.append(
        (
            "reports_by_status_all_time.csv",
            ["Status", "Count"],
            _rows_from_dicts(me_metrics["all_time"]["by_status"], "label", "count"),
        )
    )
    prov_period_rows = _rows_from_dicts(me_metrics["province_window_top"], "name", "count")
    if me_metrics.get("province_window_other_count"):
        prov_period_rows.append(["All other provinces", str(me_metrics["province_window_other_count"])])
    parts.append(("provinces_by_reports_period.csv", ["Province", "Reports"], prov_period_rows))

    prov_all_rows = _rows_from_dicts(me_metrics["province_top"], "name", "count")
    if me_metrics.get("province_other_count"):
        prov_all_rows.append(["All other provinces", str(me_metrics["province_other_count"])])
    parts.append(("provinces_by_reports_all_time.csv", ["Province", "Reports"], prov_all_rows))

    rej = me_metrics.get("rejections_window") or {}
    parts.append(
        (
            "rejections_by_reason_period.csv",
            ["Reason", "Count"],
            _rows_from_dicts(rej.get("reasons") or [], "label", "count"),
        )
    )
    parts.append(
        (
            "rejections_by_sample_type_period.csv",
            ["Sample type", "Rejected total"],
            _rows_from_dicts(rej.get("by_sample_type") or [], "label", "count"),
        )
    )

    ref = me_metrics.get("referred_window") or {}
    parts.append(
        (
            "referrals_by_test_type_period.csv",
            ["Test type", "Records", "Samples out"],
            _rows_from_dicts(ref.get("by_test_type") or [], "label", "records", "samples"),
        )
    )

    pc = me_metrics.get("pc_transport_window") or {}
    parts.append(
        (
            "non_ist_transport_period.csv",
            ["Channel", "Count"],
            [
                ["Ambulance", str(pc.get("specimens_ambulance", 0))],
                ["Alternative IP transport", str(pc.get("specimens_alternative_ip_transport", 0))],
                ["MOHCC arranged", str(pc.get("specimens_mohcc_arranged_transport", 0))],
                ["Courier", str(pc.get("specimens_courier", 0))],
                ["Other non-IST", str(pc.get("specimens_other_non_ist", 0))],
            ],
        )
    )

    delivery = me_metrics.get("delivery") or {}
    parts.append(
        (
            "specimens_by_province_period.csv",
            ["Province", "Specimens (trip rows)"],
            _rows_from_dicts(delivery.get("province_top_specimens") or [], "name", "volume"),
        )
    )

    chart = me_metrics.get("chart") or {}
    labels = chart.get("labels") or []
    trend_rows = []
    reports = chart.get("reports") or []
    samples = chart.get("samples") or []
    for i, lab in enumerate(labels):
        trend_rows.append(
            [
                str(lab),
                str(reports[i] if i < len(reports) else 0),
                str(samples[i] if i < len(samples) else 0),
            ]
        )
    parts.append(
        ("weekly_trend_period.csv", ["Week start", "Reports", "Samples"], trend_rows)
    )

    return parts


def response_parts_zip_csv(*, parts: list[tuple[str, list[str], list[list[str]]]], download_stem: str) -> HttpResponse:
    bio = BytesIO()
    with zipfile.ZipFile(bio, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, hdr, body in parts:
            zf.writestr(name, _csv_bytes(hdr, body), compress_type=zipfile.ZIP_DEFLATED)
    resp = HttpResponse(bio.getvalue(), content_type="application/zip")
    resp["Content-Disposition"] = f'attachment; filename="{filename_safe(download_stem)}.zip"'
    return resp


def response_parts_xlsx(
    *,
    parts: list[tuple[str, list[str], list[list[str]]]],
    download_stem: str,
) -> HttpResponse:
    wb = Workbook()
    wb.remove(wb.active)
    for name, hdr, body in parts:
        title = name.replace(".csv", "").replace("_", " ")[:31]
        ws = wb.create_sheet(title=title or "Sheet")
        ws.append(hdr)
        for row in body:
            ws.append(row)
    bio = BytesIO()
    wb.save(bio)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename_safe(download_stem)}.xlsx"'
    return resp
